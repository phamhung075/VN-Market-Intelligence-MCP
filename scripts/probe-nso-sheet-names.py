#!/usr/bin/env python3
"""
Probe script: NSO monthly Excel sheet names and structure.

Purpose: Diagnose sheet naming patterns in the current NSO monthly customs Excel.
Used by ops to verify sheet-name drift and confirm content-pattern identification rules.

Usage:
  ./scripts/probe-nso-sheet-names.py [--url <excel_url>] [--file <local_xlsx>] [--verbose]

Examples:
  # Fetch from NSO URL (default: June 2026)
  ./scripts/probe-nso-sheet-names.py

  # Fetch from specific URL
  ./scripts/probe-nso-sheet-names.py --url "https://nso.gov.vn/uploads/2026/07/file.xlsx"

  # Analyze local file
  ./scripts/probe-nso-sheet-names.py --file /tmp/nso-2026-07.xlsx --verbose

Output: JSON report with sheet names, headers (A1), and structure signatures.
"""

import sys
import json
import argparse
from pathlib import Path
from urllib.request import urlopen
from io import BytesIO

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def fetch_excel(url: str) -> BytesIO:
    """Fetch Excel file from URL and return as BytesIO."""
    try:
        with urlopen(url) as response:
            return BytesIO(response.read())
    except Exception as e:
        print(f"ERROR: Failed to fetch {url}: {e}", file=sys.stderr)
        sys.exit(1)


def analyze_excel(excel_bytes: BytesIO, verbose: bool = False) -> dict:
    """Analyze Excel workbook and extract sheet info."""
    try:
        wb = load_workbook(excel_bytes, data_only=True)
    except Exception as e:
        print(f"ERROR: Failed to parse Excel: {e}", file=sys.stderr)
        sys.exit(1)

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get header (A1)
        header = ws.cell(1, 1).value or ""
        if header:
            header = str(header).strip()

        # Check for keyword patterns
        keywords = []
        if "xuất khẩu" in header.lower():
            keywords.append("export")
        if "nhập khẩu" in header.lower():
            keywords.append("import")
        if "đầu tư" in header.lower() or "FDI" in header.upper():
            keywords.append("fdi")

        # Structure info
        max_row = ws.max_row
        max_col = ws.max_column

        # Look for total row markers
        total_markers = []
        for row_idx in range(1, min(15, max_row + 1)):
            a1 = ws.cell(row_idx, 1).value
            if a1:
                a1_str = str(a1).upper()
                if "TỔNG" in a1_str or "TOTAL" in a1_str:
                    total_markers.append(row_idx)

        sheet_info = {
            "name": sheet_name,
            "header_a1": header,
            "keywords": keywords,
            "dimensions": f"A1:{chr(64 + max_col)}{max_row}",
            "max_row": max_row,
            "max_col": max_col,
            "total_row_markers": total_markers,
        }

        if verbose:
            # Include first 5 rows
            first_rows = []
            for row_idx in range(1, min(6, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(6, max_col + 1)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val is None:
                        row_data.append(None)
                    else:
                        row_data.append(str(cell_val)[:50])
                first_rows.append(row_data)
            sheet_info["first_rows_verbose"] = first_rows

        sheets.append(sheet_info)

    return {
        "status": "ok",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Probe NSO monthly Excel sheet names and structure."
    )
    parser.add_argument(
        "--url",
        default="https://www.nso.gov.vn/wp-content/uploads/2026/06/02.-Bieu-T5.2026-final.xlsx",
        help="URL of NSO Excel file (default: June 2026)",
    )
    parser.add_argument("--file", help="Local path to .xlsx file (overrides --url)")
    parser.add_argument(
        "--verbose", action="store_true", help="Include first 5 rows of each sheet"
    )
    parser.add_argument(
        "--json", action="store_true", help="Output as JSON (default: human-readable)"
    )

    args = parser.parse_args()

    # Load Excel
    if args.file:
        if not Path(args.file).exists():
            print(f"ERROR: File not found: {args.file}", file=sys.stderr)
            sys.exit(1)
        excel_bytes = BytesIO(Path(args.file).read_bytes())
    else:
        excel_bytes = fetch_excel(args.url)

    # Analyze
    result = analyze_excel(excel_bytes, verbose=args.verbose)

    # Output
    if args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(f"NSO Excel Analysis ({result['sheet_count']} sheets)\n")
        for sheet in result["sheets"]:
            print(f"Sheet: {sheet['name']}")
            print(f"  Header (A1): {sheet['header_a1']}")
            print(f"  Keywords: {', '.join(sheet['keywords']) if sheet['keywords'] else '(none)'}")
            print(f"  Dimensions: {sheet['dimensions']}")
            print(f"  Total rows found at: {sheet['total_row_markers']}")
            if args.verbose and sheet.get("first_rows_verbose"):
                print(f"  First 5 rows:")
                for row_idx, row_data in enumerate(sheet["first_rows_verbose"], 1):
                    print(f"    Row {row_idx}: {row_data}")
            print()


if __name__ == "__main__":
    main()
